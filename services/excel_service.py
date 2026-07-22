from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font


class ExcelService:

    def gerar(self, origem, analise, rotas):

        workbook = Workbook()

        planilha = workbook.active
        planilha.title = "Rotas"

        titulo = Font(
            bold=True,
            size=12
        )

        planilha["A1"] = "RouteMind"
        planilha["A1"].font = titulo

        planilha["A3"] = "Origem"
        planilha["B3"] = origem

        planilha["A5"] = "Motoristas"
        planilha["B5"] = analise["motoristas"]

        planilha["A6"] = "Distância Total (km)"
        planilha["B6"] = analise["distancia_total_km"]

        planilha["A7"] = "Tempo Total (min)"
        planilha["B7"] = analise["tempo_total"]

        linha = 10

        for rota in rotas:

            planilha.cell(
                row=linha,
                column=1
            ).value = f"Motorista {rota['motorista']}"

            planilha.cell(
                row=linha,
                column=1
            ).font = titulo

            linha += 1

            planilha.cell(
                row=linha,
                column=1
            ).value = "Parada"

            planilha.cell(
                row=linha,
                column=2
            ).value = "Endereço"

            planilha.cell(
                row=linha,
                column=3
            ).value = "Bairro"

            planilha.cell(
                row=linha,
                column=4
            ).value = "Cidade"

            for coluna in range(1, 5):

                planilha.cell(
                    row=linha,
                    column=coluna
                ).font = titulo

            linha += 1

            parada = 1

            for ponto in rota["roteiro"]:

                if ponto["tipo"] != "Entrega":
                    continue

                planilha.cell(
                    row=linha,
                    column=1
                ).value = parada

                planilha.cell(
                    row=linha,
                    column=2
                ).value = ponto["endereco"]

                planilha.cell(
                    row=linha,
                    column=3
                ).value = ponto["bairro"]

                planilha.cell(
                    row=linha,
                    column=4
                ).value = ponto["cidade"]

                linha += 1
                parada += 1

            linha += 2

        arquivo = BytesIO()

        workbook.save(arquivo)

        arquivo.seek(0)

        return arquivo